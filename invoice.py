import requests
import json
import openpyxl
import os
import shutil
import glob


# Set up authentication for invoice API
invoice_client_secret = "4V9XMejsDe"

# Check if invoice token file exists
if not os.path.isfile("invoice_token.txt"):
    # If file does not exist, run invoice_token.py script to generate token
    os.system("python invoice_token.py")

# Read the access token from file
with open("invoice_token.txt", "r") as f:
    invoice_access_token = f.read().strip()

# Set the API endpoint and authentication headers for invoice API
invoice_url = 'https://api.fortnox.se/3/invoices'

invoice_headers = {
    'Authorization': f'Bearer {invoice_access_token}',
    'Client-Secret': invoice_client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Read customer data from the "Janssons kranar.txt" file
customers = {}
with open("Janssons kranar.txt", "r") as f:
    for line in f:
        name, email, phone, customer_number = [item.strip() for item in line.split(',')]
        name = name.split(': ')[1]
        email = email.split(': ')[1]
        phone = phone.split(': ')[1]
        customer_number = customer_number.split(': ')[1]

#print(customers)  # Print the customers dictionary to check if it's populated correctly

xlsx_file = "merged_janssons_kranar.xlsx"

# Check if there are any .xlsx files in the directory
xlsx_files = glob.glob(xlsx_file)

if not xlsx_files:
    print("All invoices done, for now.")
    exit()
else:
    # Load the workbook
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active

# Find the column indices for the required columns
article_number_index = None
customer_name_index = None
customer_phone_index = None
unit_index = None
delivered_quantity_index = None
price_index = None
description_index = None
unit_index = None
formname_index = None
customer_email_index = None

for i, cell in enumerate(ws[1]):
    if cell.value == "ID":
        article_number_index = i + 1
    elif cell.value == "Namn":
        customer_name_index = i + 1
    elif cell.value == "Telefonummer":
        customer_phone_index = i + 1
    elif cell.value == "Unit":
        unit_index = i + 1
    elif cell.value == "days_summarized":
        delivered_quantity_index = i + 1
    elif cell.value == "final_price":
        price_index = i + 1
    elif cell.value == "Description":
        description_index = i + 1
    elif cell.value == "formname":
        formname_index = i + 1
    elif cell.value == "Email" or cell.value == "email":
        customer_email_index = i + 1

if None in [unit_index]:
    unit_index = 'st'

# Check if all the required columns are found
if None in [article_number_index, customer_name_index, customer_phone_index, unit_index, delivered_quantity_index, price_index, description_index]:
    print("Error: Required column not found")
    #print(article_number_index, customer_name_index, customer_phone_index, unit_index, delivered_quantity_index, price_index, description_index)
    exit()

#Check if the "formname" column is found
if None in [article_number_index, customer_name_index, customer_phone_index, unit_index, delivered_quantity_index, price_index, description_index, formname_index]:
    print("Error: Required column not found")
    exit()

if formname_index is not None:
    form_name = ws.cell(row=2, column=formname_index).value
else:
    form_name = "default_form_name"

# Create a dictionary to store customer orders
customer_orders = {}

# Iterate through the rows and populate the customer_orders dictionary
for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    customer_name = str(row[customer_name_index - 1]).strip()
    customer_phone = str(row[customer_phone_index - 1]).strip()

    if customer_name == "None":
        # Look up the customer name using the phone number from the .txt file
        for key, value in customers.items():
            name, phone = key.split("_")
            if phone == customer_phone:
                customer_name = name
                break

    customer_key = f"{customer_name}_{customer_phone}"

    #article_number = row[article_number_index - 1]
    article_number = 1
    #unit = row[unit_index - 1]
    unit = unit_index
    #delivered_quantity = row[delivered_quantity_index - 1]
    delivered_quantity = 1
    price = row[price_index - 1]
    description = row[description_index - 1]

    order = {
        'ArticleNumber': article_number,
        'Unit': unit,
        'DeliveredQuantity': delivered_quantity,
        'Price': price,
        'Description': description
    }

    if customer_key in customer_orders:
        customer_orders[customer_key].append(order)
    else:
        customer_orders[customer_key] = [order]

#print(customer_orders)  # Print the customer_orders dictionary to check if it's populated correctly

# Load customer data from the text file
customer_data = {}
with open("Janssons kranar.txt", "r") as file:
    for line in file.readlines():
        if not line.strip():
            continue

        items = [item.strip() for item in line.split(',')]
        name = items[0].split(': ')[1]
        email = items[1].split(': ')[1]  # Store the email, but it won't be used in the rest of the code
        phone = items[2].split(': ')[1]
        customer_number = items[3].split(': ')[1]
        key = f"{name}_{phone}"
        customer_data[key] = customer_number

def create_invoice(customer_key, customer_info, customer_email, form_name):
    customer_name, customer_phone = customer_key.split("_")
    customer_number = get_customer_number(customer_name, customer_phone, customer_data)

    if not customer_number:
        print(f"Error: Customer {customer_name} with phone number {customer_phone} not found in the text file")
        
        # Write the error details to the text file
        error_file_name = f"{form_name}_customer_error.txt"
        with open(error_file_name, "a") as error_file:
            error_file.write(f"Customer: {customer_name}\n")
            error_file.write(f"Phone: {customer_phone}\n")
            error_file.write(f"Email: {customer_email}\n")
            error_file.write("Error: Customer not found in the text file\n")
            error_file.write("\n")  # Add an empty line to separate the error details of different customers
        
        return

    # Create an invoice with the customer number and order data
    invoice_data = {
        "Invoice": {
            "CustomerNumber": customer_number,
            "InvoiceRows": customer_info
        }
    }

    response = requests.post(invoice_url, headers=invoice_headers, json=invoice_data)

    if response.status_code == 201:
        invoice_number = response.json()["Invoice"]["DocumentNumber"]
        print(f"Created invoice {invoice_number} for customer {customer_name} with phone number {customer_phone}")
    elif response.status_code == 401:  # Unauthorized
        print("Access token expired. Refreshing the access token...")
        os.system("python invoice_token.py")
        with open("invoice_token.txt", "r") as f:
            invoice_access_token = f.read().strip()
        invoice_headers['Authorization'] = f'Bearer {invoice_access_token}'
        create_invoice(customer_key, customer_info, customer_email, form_name)  # Retry creating the invoice after refreshing the access token
    else:
        print(f"Error: {response.content}")
        print(order)

def get_customer_number(name, phone, customer_data):
    key = f"{name}_{phone}"
    
    # Check if the key is in the customer_data
    if key in customer_data:
        return customer_data[key]
    
    # If not found, check if adding a "0" at the beginning helps
    if phone[0] != '0':
        key_with_zero = f"{name}_0{phone}"
        if key_with_zero in customer_data:
            return customer_data[key_with_zero]
    
    print(f"Error: Customer {name} with phone number {phone} not found in the text file")
    return None
    
#checking error and email before
def get_email_by_phone(phone, ws, phone_index, email_index, customers):
    for key, value in customers.items():
        name, phone_from_txt = key.split("_")
        if phone_from_txt == phone or phone_from_txt == f"0{phone}":
            return value[1]

    for row_number, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        phone_from_xlsx = str(row[phone_index - 1]).strip()
        if phone_from_xlsx == phone or phone_from_xlsx == f"0{phone}":
            return ws.cell(row=row_number, column=email_index).value
    return None
   
# Create the invoices
for customer_key, customer_info in customer_orders.items():
    customer_email = customers.get(customer_key, [None, None])[1]  # Get the email from the customers dictionary

    if customer_email is None:
        customer_email = get_email_by_phone(customer_key.split("_")[1], ws, customer_phone_index, customer_email_index, customers)

    if customer_email is None:
        print(f"Error: Email not found for customer with phone number {customer_key.split('_')[1]}")
        continue

    create_invoice(customer_key, customer_info, customer_email, form_name)

# Move the used .xlsx file to the "invoice_done" folder
source = xlsx_file
destination_folder = "invoice_done"

if not os.path.exists(destination_folder):
    os.makedirs(destination_folder)

destination = os.path.join(destination_folder, source)
shutil.move(source, destination)
