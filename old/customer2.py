import requests
import json
import openpyxl
import os

# Set up authentication
client_secret = "4V9XMejsDe"

# Check if customer token file exists
if not os.path.isfile("customer_token.txt"):
    # If file does not exist, run customer_token.py script to generate token
    os.system("python customer_token.py")

# Read the access token from file
with open("customer_token.txt", "r") as f:
    access_token = f.read().strip()

# Set the API endpoint and authentication headers
customer_url = 'https://api.fortnox.se/3/customers'

wb = openpyxl.load_workbook('test2.xlsx')
ws = wb.active

headers = {
    'Authorization': f'Bearer {access_token}',
    'Client-Secret': client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

#print([cell.value for cell in ws[1]])

# Set the query parameter to search for the customer by the name
namn_column_index = None
email_column_index = None
address_column_index = None
zipcode_column_index = None
city_column_index = None

for i, cell in enumerate(ws[1]):
    if cell.value == "Namn":
        namn_column_index = i + 1
    elif cell.value == "Email":
        email_column_index = i + 1
    elif cell.value == "Adress":
        address_column_index = i + 1
    elif cell.value == "Postnummer":
        zipcode_column_index = i + 1
    elif cell.value == "Stad":
        city_column_index = i + 1

# Set the query parameter to search for the customer by the name
if namn_column_index is not None:
    #print("namn_column_index found:", namn_column_index)
    params = {
        'name': ws.cell(row=2, column=namn_column_index).value
    }

if namn_column_index is None:
    print("Column is none, error.")
    exit()

# Send a GET request to the Fortnox API to search for the customer by name
response = requests.get(customer_url, headers=headers, params=params)

# Check if access token has expired
if response.status_code == 401:
    # If access token has expired, run customer_token.py script to generate new token
    os.system("python customer_token.py")
    with open("customer_token.txt", "r") as f:
        access_token = f.read().strip()
    headers['Authorization'] = f'Bearer {access_token}'
    # Send a new GET request to search for the customer by name
    response = requests.get(customer_url, headers=headers, params=params)

# Check the response status code and print the response content
if response.status_code == 200:
    # Parse the JSON response data
    response_data = response.json()
    # If the customer exists, use the existing customer number
    if response_data['MetaInformation']['@TotalResources'] > 0:
        customer_number = response_data['Customers'][0]['Name']
    # If the customer does not exist, create a new customer
    else:
        # Create the customer data payload
        customer_data = {
                'Customer': {
                    'Name': ws.cell(row=2, column=namn_column_index).value,
                    'Email': ws.cell(row=2, column=email_column_index).value,
                    'Address1': ws.cell(row=2, column=address_column_index).value,
                    'ZipCode': ws.cell(row=2, column=zipcode_column_index).value,
                    'City': ws.cell(row=2, column=city_column_index).value
            }
        }
        # Send a POST request to the Fortnox API to create a new customer
        response = requests.post(customer_url, headers=headers, json=customer_data)
        # Parse the JSON response data
        response_data = response.json()
        # Get the customer number from the response data
        customer_number = response_data['Customer']['Name']

# Check the response status code and print the response content
if response.status_code == 200 or response.status_code == 201:
    print('Assignment successfully done! Created or customer already existed!')
else:
    print('Error: {}'.format(response.content))
