import requests
import json
import openpyxl
import os

# Set up authentication
client_secret = "4V9XMejsDe"

# Check if invoice token file exists
if not os.path.isfile("invoice_token.txt"):
    # If file does not exist, run invoice_token.py script to generate token
    os.system("python invoice_token.py")

# Read the access token from file
with open("invoice_token.txt", "r") as f:
    access_token = f.read().strip()

# Set the API endpoint and authentication headers
invoice_url = 'https://api.fortnox.se/3/invoices'

headers = {
    'Authorization': f'Bearer {access_token}',
    'Client-Secret': client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Check if access token has expired
response = requests.get(invoice_url, headers=headers)
if response.status_code == 401:
    # If access token has expired, run invoice_token.py script to generate new token
    os.system("python invoice_token.py")
    with open("invoice_token.txt", "r") as f:
        access_token = f.read().strip()
    headers['Authorization'] = f'Bearer {access_token}'

# Read the customer name, customer number, and invoice data from Excel
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active
customer_name = ws.cell(row=2, column=1).value
customer_number = ws.cell(row=2, column=2).value
invoice_data = {
    'Invoice': {
        'CustomerName': customer_name,
        'CustomerNumber': customer_number,
        'InvoiceRows': [
            {
                'ArticleNumber': ws.cell(row=i, column=1).value,
                'Quantity': ws.cell(row=i, column=4).value,
                'Price': ws.cell(row=i, column=5).value
            }
            for i in range(3, ws.max_row + 1)
        ]
    }
}

# Send a POST request to the Fortnox API to create a new invoice
response = requests.post(invoice_url, headers=headers, json=invoice_data)

# Check if access token has expired
if response.status_code == 401:
    # If access token has expired, run invoice_token.py script to generate new token
    os.system("python invoice_token.py")
    with open("invoice_token.txt", "r") as f:
        access_token = f.read().strip()
    headers['Authorization'] = f'Bearer {access_token}'
    # Send a new POST request to create a new invoice
    response = requests.post(invoice_url, headers=headers, json=invoice_data)

# Check the response status code and print the response content
if response.status_code == 200 or response.status_code == 201:
    print('Invoice successfully created!')
else:
    print('Error: {}'.format(response.content))
