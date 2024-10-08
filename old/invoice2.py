import requests
import json
import openpyxl
import os
import shutil

# Set up authentication for invoice API
invoice_client_secret = "4V9XMejsDe"

# Check if invoice token file exists
if not os.path.isfile("invoice_token.txt"):
    # If file does not exist, run invoice_token.py script to generate token
    os.system("python invoice_token.py")

# Read the access token from file
with open("invoice_token.txt", "r") as f:
    invoice_access_token = f.read().strip()

# Set the API endpoint and authentication headers for invoice API
invoice_url = 'https://api.fortnox.se/3/invoices'

invoice_headers = {
    'Authorization': f'Bearer {invoice_access_token}',
    'Client-Secret': invoice_client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Check if access token for invoice API has expired
response = requests.get(invoice_url, headers=invoice_headers)
if response.status_code == 401:
    # If access token has expired, run invoice_token.py script to generate new token
    os.system("python invoice_token.py")
    with open("invoice_token.txt", "r") as f:
        invoice_access_token = f.read().strip()
    invoice_headers['Authorization'] = f'Bearer {invoice_access_token}'

# Set up authentication for customer API
customer_client_secret = "KSnpxgS5na"

# Check if customer token file exists
if not os.path.isfile("customer_token.txt"):
    # If file does not exist, run customer_token.py script to generate token
    os.system("python customer_token.py")

# Read the access token from file
with open("customer_token.txt", "r") as f:
    customer_access_token = f.read().strip()

# Set the API endpoint and authentication headers for customer API
customer_url = 'https://api.fortnox.se/3/customers'

customer_headers = {
    'Authorization': f'Bearer {customer_access_token}',
    'Client-Secret': customer_client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Check if access token for customer API has expired
response = requests.get(customer_url, headers=customer_headers)
if response.status_code == 401:
    # If access token has expired, run customer_token.py script to generate new token
    os.system("python customer_token.py")
    with open("customer_token.txt", "r") as f:
        customer_access_token = f.read().strip()
    customer_headers['Authorization'] = f'Bearer {customer_access_token}'

# Read the invoice data from Excel
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active

# Find the column indices for the required columns
app_quantity_index = None
final_price_index = None
article_number_index = None
customer_name_index = None
customer_phone_index = None
customer_number_index = None

for i, cell in enumerate(ws[1]):
    if cell.value == "app_quantity_1":
        app_quantity_index = i + 1
    elif cell.value == "final_price":
        final_price_index = i + 1
    elif cell.value == "ID":
        article_number_index = i + 1
    elif cell.value == "Namn":
        customer_name_index = i + 1
    elif cell.value == "Telefonummer":
        customer_phone_index = i + 1
    elif cell.value == "Personummer":
        customer_number_index = i + 1

# Check if all the required columns are found
if None in [app_quantity_index, final_price_index, article_number_index, customer_name_index, customer_phone_index, customer_number_index]:
    print("Error: Required column not found")
    exit()

# Create the invoices
# Create the invoices
for row in ws.iter_rows(min_row=2, values_only=True):
    customer_name = row[customer_name_index - 1]
    customer_phone = row[customer_phone_index - 1]
    customer_number = row[customer_number_index - 1]

    # Check if customer name and phone number are provided
    if customer_name is None or customer_phone is None:
        print(f'Error: Customer name or phone number missing in row {row}')
        continue

    # Read the invoice data from Excel
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb.active
    # Extract the invoice data from the Excel sheet
    invoice_data = {
        'Invoice': {
            'name': customer_name,
            'InvoiceRows': []
        }
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or row[1] is None or row[3] is None or row[4] is None:
            continue
        article_number = row[0]
        quantity = row[3]
        price = row[4]

        # Add the invoice row to the invoice data
        invoice_row = {
            'ArticleNumber': article_number,
            'Quantity': quantity,
            'Price': price
        }
        invoice_data['Invoice']['InvoiceRows'].append(invoice_row)

    # Send a POST request to the Fortnox API to create a new invoice
    response = requests.post(invoice_url,  headers=invoice_headers, json=invoice_data)

    # Check if access token has expired
    if response.status_code == 401:
        # If access token has expired, generate a new token and retry
        with open("invoice_token.txt", "r") as f:
            access_token = f.read().strip()
        headers['Authorization'] = f'Bearer {access_token}'
        response = requests.post(invoice_url, headers=headers, json=invoice_data)

    # Check the response status code and print the response content
    if response.status_code == 200 or response.status_code == 201:
        print('Invoice successfully created!')
    else:
        print('Error: {}'.format(response.content))

    # Move the Excel file to the "done" folder
    os.makedirs("done", exist_ok=True)
    os.replace("test.xlsx", os.path.join("done", "test.xlsx"))
