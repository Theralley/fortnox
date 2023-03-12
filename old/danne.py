import requests
import json
import pandas as pd
import webbrowser
import time
import base64
import openpyxl
from urllib.parse import urlparse, parse_qs

# Set up authentication
client_secret = "4V9XMejsDe"
client_id = "MYslrKdiUz4o"
auth_url = "https://apps.fortnox.se/oauth-v1/auth"

auth_params = {
    "client_id": client_id,
    "redirect_uri": "https://mysite.org/activation",
    #"scope": "customer",
    "scope": "invoice",
    "state": "somestate123",
    "access_type": "offline",
    "response_type": "code",
    "account_type": "service"
}

# Authorize and get authorization code
auth_response = requests.get(auth_url, params=auth_params)
if auth_response.status_code != 200:
    print("Error authorizing with Fortnox API")
    print(auth_response.text)
    exit()

webbrowser.open(auth_response.url)
time.sleep(10)  # Wait for user to log in
redirect_url = input("Enter the URL you were redirected to: ")
query_params = parse_qs(urlparse(redirect_url).query)
auth_code = query_params["code"][0]

token_url = "https://api.fortnox.se/oauth-v1/token"
redirect_uri = "https://mysite.org/activation"

# Set the request headers
headers = {
    'Content-Type': 'application/x-www-form-urlencoded',
    'Authorization': 'Basic {}'.format(base64.b64encode('{}:{}'.format(client_id, client_secret).encode()).decode())
}

# Set the request body
body = {
    'grant_type': 'authorization_code',
    'client_id': client_id,
    'redirect_uri': 'https://mysite.org/activation',
    'code': auth_code
}

# Send a POST request to the token URL using the requests library
token_response = requests.post(token_url, headers=headers, data=body)

# Parse the token response
token_data = token_response.json()
access_token = token_data['access_token']
refresh_token = token_data['refresh_token']
expires_in = token_data['expires_in']

# Set the API endpoint and authentication headers
invoices_url = 'https://api.fortnox.se/3/invoices'

# Set the API endpoint and authentication headers
customer_url = 'https://api.fortnox.se/3/customers'

wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active

headers = {
    'Authorization': f'Bearer {access_token}',
    'Client-Secret': client_secret,
    'Content-Type': 'application/json',
    'Accept':"application/json"
}

# Set the query parameter to search for the customer by the name
params = {
    'name': ws.cell(row=2, column=1).value
}

# Send a GET request to the Fortnox API to search for the customer by name
response = requests.get(invoices_url, headers=headers, params=params)
#esponse.status_code = 1
# Check the response status code and print the response content
if response.status_code == 200:
    # Parse the JSON response data
    response_data = response.json()
    # If the customer exists, use the existing customer number
    if response_data['MetaInformation']['@TotalResources'] > 0:
        customer_number = response_data['Customers'][0]['CustomerNumber']
    # If the customer does not exist, create a new customer
    else:
        # Create the customer data payload
        customer_data = {
            'Customer': {
                'Name': ws.cell(row=2, column=1).value,
                'Email': ws.cell(row=2, column=3).value,
                'Address1': ws.cell(row=2, column=6).value,
                'ZipCode': ws.cell(row=2, column=5).value,
                'City': ws.cell(row=2, column=4).value
            }
        }
        # Send a POST request to the Fortnox API to create a new customer
        response = requests.post(customer_url, headers=headers, json=customer_data)
        # Parse the JSON response data
        response_data = response.json()
        # Get the customer number from the response data
        customer_number = response_data['Customer']['CustomerNumber']

else:
    print('Error: {}'.format(response.content))
    exit()

response = requests.get(invoices_url, headers=headers, params=params)

invoice_data = {
    'Invoice': {
        'CustomerName': ws.cell(row=2, column=1).value,
        'CustomerNumber': customer_number,
        'InvoiceRows': [
            {
                'ArticleNumber': ws.cell(row=i, column=1).value,
                'Quantity': ws.cell(row=i, column=4).value,
                'Price': ws.cell(row=i, column=5).value
            }
            for i in range(3, ws.max_row + 1)
        ]
    }
}

# Send the invoice data to Fortnox using the requests library
response = requests.post(invoices_url, headers=headers, json=invoice_data)

# Check the response status code and print the response content
if response.status_code == 201:
    print('Assignment successfully done!')
else:
    print('Error: {}'.format(response.content))
