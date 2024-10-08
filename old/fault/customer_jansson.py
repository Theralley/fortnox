import requests
import json
import openpyxl
import os

# Set up authentication
client_secret = "4V9XMejsDe"

# Check if customer token file exists
if not os.path.isfile("customer_token.txt"):
    # If file does not exist, run customer_token.py script to generate token
    os.system("python customer_token.py")

# Read the access token from file
with open("customer_token.txt", "r") as f:
    access_token = f.read().strip()


# Set the API endpoint and authentication headers
customer_url = 'https://api.fortnox.se/3/customers'

directory = os.getcwd()

# Find all xlsx files in the directory containing "jansson" in the file name
xlsx_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') and 'janssons_kranar_03_15_23' in f]

# Loop through xlsx files and load the workbook
for xlsx_file in xlsx_files:
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    # Rest of your code for processing the workbook goes her

headers = {
    'Authorization': f'Bearer {access_token}',
    'Client-Secret': client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Set the query parameter to search for the customer by the name
namn_column_index = None
email_column_index = None
address_column_index = None
zipcode_column_index = None
city_column_index = None
org_personnummer_index = None
formname_index = None
customer_phone_index = None

for i, cell in enumerate(ws[1]):
    if cell.value == "Namn":
        namn_column_index = i + 1
    elif cell.value == "Email":
        email_column_index = i + 1
    elif cell.value == "Adress":
        address_column_index = i + 1
    elif cell.value == "Postnummer":
        zipcode_column_index = i + 1
    elif cell.value == "Stad":
        city_column_index = i + 1
    elif cell.value == "Personummer":
        org_personnummer_index = i + 1
    elif cell.value == "formname":
        formname_index = i + 1
    elif cell.value == "Telefonummer":
        customer_phone_index = i + 1

# Set the query parameter to search for the customer by the name
if namn_column_index is not None:
    params = {
        'name': ws.cell(row=2, column=namn_column_index).value
    }

if namn_column_index is None:
    print("Column is none, error.")
    exit()

customer_data = {
    'Customer': {}
}

for row in ws.iter_rows(min_row=2, values_only=True):
    if namn_column_index is not None and row[namn_column_index-1]:
        params = {
            'name': row[namn_column_index-1]
        }

        # Define email
        if email_column_index is not None:
            email = row[email_column_index-1]
        else:
            email = None

        # Define address
        if address_column_index is not None:
            address = row[address_column_index-1]
        else:
            address = None

        # Define zipcode
        if zipcode_column_index is not None:
            zipcode = row[zipcode_column_index-1]
        else:
            zipcode = None

        # Define city
        if city_column_index is not None:
            city = row[city_column_index-1]
        else:
            city = None

        # Define Costumer Number
        if org_personnummer_index is not None:
            org_personnummer = row[org_personnummer_index-1]
        else:
            org_personnummer = None

        # Define Costumer Phone
        if customer_phone_index is not None:
            customer_phone = row[customer_phone_index-1]
        else:
            customer_phone = None

        # Define formname
        if formname_index is not None:
            formname = row[formname_index-1]
            filename = f"{formname}.txt"
        else:
            formname = None


        # Define customer_data
        customer_data = {
            'Customer': {
                'Name': row[namn_column_index-1],
                'Email': email,
                'Address1': address,
                'ZipCode': zipcode,
                'City': city,
                'OrganisationNumber': org_personnummer,
                'Phone1': customer_phone
            }
        }

        # Send a GET request to the Fortnox API to search for the customer by name
        response = requests.get(customer_url, headers=headers, params=params)

        # Check if access token has expired
        if response.status_code == 401:
            # If access token has expired, run customer_token.py script to generate new token
            os.system("python customer_token.py")
            with open("customer_token.txt", "r") as f:
                access_token = f.read().strip()
            headers['Authorization'] = f'Bearer {access_token}'
            # Send a new GET request to search for the customer by name
            response = requests.get(customer_url, headers=headers, params=params)

        # Check the response status code and print the response content
        if response.status_code == 200:
            # Parse the JSON response data
            response_data = response.json()
            # If the customer exists, use the existing customer number
            if response_data['MetaInformation']['@TotalResources'] > 0:
                customer_number = response_data['Customers'][0]['CustomerNumber']
                print(f'Customer {row[namn_column_index-1]} already exists! Customer number is: {customer_number}')
                with open(filename, "a") as f:
                    f.write(f"Name: {row[namn_column_index-1]}, ")
                    f.write(f"Phone number: {customer_phone}, ")
                    f.write(f"Customer number: {customer_number}\n")
            # If the customer does not exist, create a new customer
            else:
                # Send a POST request to the Fortnox API to create a new customer
                response = requests.post(customer_url, headers=headers, json=customer_data)
                # Parse the JSON response data
                response_data = response.json()
                # Get the customer number from the response data
                customer_number = response_data['Customer']['CustomerNumber']

                # Check the response status code and print the response content
                if response.status_code == 200 or response.status_code == 201:
                    print(f'Customer {row[namn_column_index-1]} assignment successful! Customer number is: {customer_number})')

                    with open(filename, "w") as f:
                        f.write(f"Name: {row[namn_column_index-1]}, ")
                        f.write(f"Phone number: {customer_phone}, ")
                        f.write(f"Customer number: {customer_number}\n")
                else:
                    print(f'Error: {response.content}')
