# Create the invoices
import requests
import json
import openpyxl
import os

# Set up authentication for invoice API
invoice_client_secret = "4V9XMejsDe"

# Check if customer token file exists
if not os.path.isfile("customer_token.txt"):
    # If file does not exist, run customer_token.py script to generate token
    os.system("python customer_token.py")

# Read the access token from file
with open("customer_token.txt", "r") as f:
    customer_access_token = f.read().strip()

# Check if invoice token file exists
if not os.path.isfile("invoice_token.txt"):
    # If file does not exist, run invoice_token.py script to generate token
    os.system("python invoice_token.py")

# Read the invoice token from file
with open("invoice_token.txt", "r") as f:
    invoice_access_token = f.read().strip()

# Set the API endpoint and authentication headers for invoice API
invoice_url = 'https://api.fortnox.se/3/invoices'

invoice_headers = {
    'Authorization': f'Bearer {invoice_access_token}',
    'Client-Secret': invoice_client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Set up authentication for customer API
customer_client_secret = "KSnpxgS5na"

# Set the API endpoint and authentication headers for customer API
customer_url = 'https://api.fortnox.se/3/customers'

customer_headers = {
    'Authorization': f'Bearer {customer_access_token}',
    'Client-Secret': customer_client_secret,
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

# Set directory
directory = os.getcwd()

# Find all xlsx files in the directory containing "jansson" in the file name
xlsx_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') and 'janssons_kranar_03_15_23' in f]

# Loop through xlsx files and load the workbook
for xlsx_file in xlsx_files:
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    # Rest of your code for processing the workbook goes here...

# Find the column indices for the required columns
app_quantity_index = None
final_price_index = None
article_number_index = None
customer_name_index = None
customer_phone_index = None
customer_number_index = None
hire_index = None
description_index = None

for i, cell in enumerate(ws[1]):
    if cell.value.startswith("app_quantity"):
        app_quantity_index = i + 1
    elif cell.value.startswith("final_price"):
        final_price_index = i + 1
    elif cell.value == "ID":
        article_number_index = i + 1
    elif cell.value == "Namn":
        customer_name_index = i + 1
    elif cell.value == "Telefonummer":
        customer_phone_index = i + 1
    elif cell.value == "Personummer":
        customer_number_index = i + 1
    elif cell.value == "formname":
        formname_index = i + 1
    elif cell.value.startswith("app_servic"):
        description_index = i + 1
    elif cell.value == "days_hire":
        hire_index = i + 1

# Check if all the required columns are found
if None in [app_quantity_index, final_price_index, article_number_index, customer_name_index, customer_phone_index, customer_number_index]:
    print("Error: Required column not found")
    print(app_quantity_index, final_price_index, article_number_index, customer_name_index, customer_phone_index, customer_number_index)
    exit()

if hire_index is None:
    print("Error: 'days_hire' column not found")
    exit()

if None in [app_quantity_index, final_price_index, article_number_index, customer_name_index, customer_phone_index, customer_number_index, description_index]:
    print("Error: Required column not found")
    print(app_quantity_index, final_price_index, article_number_index, customer_name_index, customer_phone_index, customer_number_index, description_index)
    exit()

# Create the invoices
for row in ws.iter_rows(min_row=2, values_only=True):
    customer_name = row[customer_name_index - 1]
    customer_phone = row[customer_phone_index - 1]
    quantity = row[app_quantity_index - 1]
    article_number =row[article_number_index - 1]
    formname = row[formname_index - 1]
    price = row[final_price_index - 1]
    description = row[description_index - 1]
    customer_number = row[customer_number_index - 1]
    hire = row[hire_index - 1]

    # Read the customer data from the text file
    filename = f"{formname}.txt"
    with open(filename) as f:
        customer_data = f.readlines()

    # Search for the customer in the customers text file by name and phone number
    with open(f"{formname}.txt", "r") as f:
        for line in f:
            if customer_name in line and str(customer_phone) in line:
                customer_number = line.split("Customer number: ")[1].strip()
                break

    if customer_number is None:
        print(f'Error: Customer {customer_name} with phone number {customer_phone} not found')
        continue

    # Extract the invoice data from the Excel sheet
    invoice_data = {
        'Invoice': {
            'CustomerNumber': customer_number,
            'InvoiceRows': []
        }
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != article_number:
            continue

        if row[1] is None or row[3] is None or row[4] is None:
            continue

        #make description
        description = "Easymarine - Hyra av " + str(description) + " i totalt " + str(hire) + " dag(ar)"
        #print(description)

        # Add the invoice row to the invoice data
        invoice_row = {
            'ArticleNumber': 1, #define article_number if needed
            'Unit': 'st', # add the unit of measurement
            'DeliveredQuantity': quantity, # should be DeliveredQuantity instead of Quantity
            'Price': price,
            'Description' : description
        }
        invoice_data['Invoice']['InvoiceRows'].append(invoice_row)

    # Send a POST request to the Fortnox API to create a new invoice
    response = requests.post(invoice_url, headers=invoice_headers, json=invoice_data)

    # Check if access token has expired
    if response.status_code == 401:
        # If access token has expired, generate a new token and retry
        with open("invoice_token.txt", "r") as f:
            access_token = f.read().strip()
        invoice_headers['Authorization'] = f'Bearer {access_token}'
        response = requests.post(invoice_url, headers=invoice_headers, json=invoice_data)

        # Update the customer number in the Excel sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            customer_name = row[customer_name_index - 1]
            customer_phone = str(row[customer_phone_index - 1])  # convert to string
            if customer_name is None or customer_phone is None:
                print("Error: Missing customer name or phone number")
                continue
            # Search for the customer in the customers text file by name and phone number
            with open(f"{formname}.txt", "r") as f:
                for line in f:
                    if customer_name is not None and customer_phone is not None and customer_name in line and str(customer_phone) in line:
                        customer_number = line.split("Customer number: ")[1].strip()
                        break
                else:
                    print(f'Error: Customer {customer_name} with phone number {customer_phone} not found')
                    continue

    # Check if the retry was successful
    if response.status_code == 200:
        print('Invoice successfully created for '+ customer_name +'!')
        response_data = response.json()
        print(response_data.keys()) # print the keys present in the dictionary
        try:
            invoice_number = response_data['Invoice']['DocumentNumber'] # use the correct key for the invoice number
        except KeyError:
            print(f"Error: {response.content}")
            continue

    if response.content[2] == 109:
        os.system("python invoice_token.py")

    else:
        print(f'Error: {response.content}')
        print('Error hos: '+ customer_name, customer_phone, customer_number, quantity, price, description)
        # Save the updated Excel sheet
        wb.save(xlsx_file)

else:
    print(f'Error: {response.content}')
