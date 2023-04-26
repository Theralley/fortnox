import openpyxl
from collections import Counter
from openpyxl.styles import Font
import os
import shutil

def xlsx_checker(file_name):
    # Load the workbook and select the first sheet
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Find columns with names app_service_1, app_service_2, app_service_3, etc.
    app_service_columns = []
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value and col[0].value.startswith("app_service_"):
            app_service_columns.append(col)

    # Create a set to store unique service names
    service_names = set()

    # Find unique service names
    for row in range(2, ws.max_row + 1):
        for col in app_service_columns:
            cell_value = col[row - 1].value
            if cell_value:
                service_names.add(cell_value)

    # Add columns for each unique service and a "Description" column
    bold_font = Font(bold=True)
    column_mapping = {}
    for idx, service in enumerate(sorted(service_names), start=ws.max_column + 1):
        ws.cell(row=1, column=idx).value = service
        ws.cell(row=1, column=idx).font = bold_font
        column_mapping[service] = idx
    ws.cell(row=1, column=ws.max_column + 1).value = "Description"
    ws.cell(row=1, column=ws.max_column).font = bold_font

    # Add "days_summarized" column
    ws.cell(row=1, column=ws.max_column + 1).value = "days_summarized"
    ws.cell(row=1, column=ws.max_column).font = bold_font

    # Iterate through rows and summarize the app_service_* columns
    for row in range(2, ws.max_row + 1):
        summary = []
        for col in app_service_columns:
            cell_value = col[row - 1].value
            if cell_value:
                summary.append(cell_value)

        # Count occurrences of each value
        count_summary = Counter(summary)

        # Add the count to the corresponding row in the correct column and prepare the description
        description_items = []
        days_sum = 0
        for key, value in count_summary.items():
            ws.cell(row=row, column=column_mapping[key]).value = value
            description_items.append(f"{key} {value} dagar")
            days_sum += value

        # Set the description in the "Description" column
        ws.cell(row=row, column=ws.max_column - 1).value = ', '.join(description_items)

        # Set the sum of rented days in the "days_summarized" column
        ws.cell(row=row, column=ws.max_column).value = days_sum

    # Save the modified workbook to the same file
    wb.save(file_name)

# Use the current working directory
folder_path = os.getcwd()

# Iterate through all files in the folder
for file in os.listdir(folder_path):
    # Check if the file has a .xlsx extension
    if file.endswith("merged_janssons_kranar.xlsx"):
        file_path = os.path.join(folder_path, file)
        xlsx_checker(file_path)

        # Move the processed file up one folder
        parent_folder = os.path.dirname(folder_path)
        new_file_path = os.path.join(parent_folder, file)
        shutil.move(file_path, new_file_path)
